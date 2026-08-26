from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from io import BytesIO
import os
import tempfile

from app.database import get_db

router = APIRouter()


# ---------- 导入 ----------
@router.post("/import/teachers")
async def import_teachers(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """批量导入教师（Excel）
    - 工号唯一，重复则跳过
    - 院系匹配：按院系名称
    """
    from openpyxl import load_workbook
    from app.models.teacher import Teacher
    from app.models.department import Department

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件")

    # 加载院系缓存
    depts = db.query(Department).all()
    dept_by_name = {d.name: d for d in depts}

    # 读取Excel
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    # 解析表头
    headers = {}
    header_row = 1
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col).value or "").strip()
        if val:
            headers[val] = col

    required_headers = ["工号", "姓名"]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要列：{', '.join(missing)}")

    def _get_val(row, col_name):
        col = headers.get(col_name)
        if not col:
            return None
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        return str(val).strip()

    def _parse_int(val, default=None):
        if val is None or val == "":
            return default
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default

    def _parse_bool(val, default=True):
        if val is None or val == "":
            return default
        s = str(val).strip()
        if s in ("是", "Y", "y", "true", "True", "1", "需要"):
            return True
        if s in ("否", "N", "n", "false", "False", "0", "不需要"):
            return False
        return default

    success_count = 0
    fail_count = 0
    skip_count = 0
    failures = []

    for row in range(header_row + 1, ws.max_row + 1):
        teacher_no = _get_val(row, "工号")
        name = _get_val(row, "姓名")

        # 跳过空行
        if not teacher_no and not name:
            continue

        row_errors = []
        if not teacher_no:
            row_errors.append("工号不能为空")
        if not name:
            row_errors.append("姓名不能为空")

        # 查重
        if teacher_no and db.query(Teacher).filter(Teacher.teacher_no == teacher_no).first():
            skip_count += 1
            continue

        # 院系匹配
        dept_name = _get_val(row, "院系")
        department_id = None
        if dept_name:
            dept = dept_by_name.get(dept_name)
            if dept:
                department_id = dept.id
            else:
                row_errors.append(f"院系「{dept_name}」不存在")

        if row_errors:
            fail_count += 1
            failures.append({"row": row, "name": name or teacher_no, "errors": row_errors})
            continue

        teacher = Teacher(
            teacher_no=teacher_no,
            name=name,
            gender=_get_val(row, "性别") or None,
            title=_get_val(row, "职称") or None,
            department_id=department_id,
            max_hours_per_day=_parse_int(_get_val(row, "日最大课时"), 6),
            max_hours_per_week=_parse_int(_get_val(row, "周最大课时"), 20),
            max_consecutive_hours=_parse_int(_get_val(row, "最大连堂"), 4),
            need_noon_break=_parse_bool(_get_val(row, "午休需求"), True),
            phone=_get_val(row, "电话") or None,
            email=_get_val(row, "邮箱") or None,
            status=1,
        )
        db.add(teacher)
        success_count += 1

    if success_count > 0:
        db.commit()

    return {
        "message": f"导入完成：成功 {success_count} 条，跳过 {skip_count} 条，失败 {fail_count} 条",
        "success_count": success_count,
        "skip_count": skip_count,
        "fail_count": fail_count,
        "failures": failures[:50],
    }


@router.post("/import/classrooms")
async def import_classrooms(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """批量导入教室（Excel）
    - 教室编号唯一，重复则跳过
    """
    from openpyxl import load_workbook
    from app.models.classroom import Classroom

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件")

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    headers = {}
    header_row = 1
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col).value or "").strip()
        if val:
            headers[val] = col

    required_headers = ["教室编号", "教学楼", "房间号"]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要列：{', '.join(missing)}")

    def _get_val(row, col_name):
        col = headers.get(col_name)
        if not col:
            return None
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        return str(val).strip()

    def _parse_int(val, default=0):
        if val is None or val == "":
            return default
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default

    type_map = {
        "普通教室": "normal", "普通": "normal", "normal": "normal",
        "多媒体教室": "multimedia", "多媒体": "multimedia", "multimedia": "multimedia",
        "实验室": "lab", "lab": "lab",
        "计算机房": "computer", "机房": "computer", "computer": "computer",
        "艺术教室": "arts", "艺术": "arts", "arts": "arts",
    }

    success_count = 0
    fail_count = 0
    skip_count = 0
    failures = []

    for row in range(header_row + 1, ws.max_row + 1):
        room_no = _get_val(row, "教室编号")
        building = _get_val(row, "教学楼")
        room_number = _get_val(row, "房间号")

        if not room_no and not building and not room_number:
            continue

        row_errors = []
        if not room_no:
            row_errors.append("教室编号不能为空")
        if not building:
            row_errors.append("教学楼不能为空")
        if not room_number:
            row_errors.append("房间号不能为空")

        if room_no and db.query(Classroom).filter(Classroom.room_no == room_no).first():
            skip_count += 1
            continue

        if row_errors:
            fail_count += 1
            failures.append({"row": row, "room_no": room_no or "", "errors": row_errors})
            continue

        type_raw = _get_val(row, "类型") or "普通教室"
        classroom_type = type_map.get(type_raw, "normal")

        classroom = Classroom(
            room_no=room_no,
            building=building,
            room_number=room_number,
            capacity=_parse_int(_get_val(row, "容量"), 50),
            classroom_type=classroom_type,
            equipment=_get_val(row, "设备") or None,
            campus=_get_val(row, "校区") or None,
            status=1,
        )
        db.add(classroom)
        success_count += 1

    if success_count > 0:
        db.commit()

    return {
        "message": f"导入完成：成功 {success_count} 条，跳过 {skip_count} 条，失败 {fail_count} 条",
        "success_count": success_count,
        "skip_count": skip_count,
        "fail_count": fail_count,
        "failures": failures[:50],
    }


@router.post("/import/courses")
async def import_courses(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """批量导入课程（Excel）
    - 课程代码唯一，重复则跳过
    - 院系匹配：按院系名称
    """
    from openpyxl import load_workbook
    from app.models.course import Course
    from app.models.department import Department
    from decimal import Decimal

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件")

    depts = db.query(Department).all()
    dept_by_name = {d.name: d for d in depts}

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    headers = {}
    header_row = 1
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col).value or "").strip()
        if val:
            headers[val] = col

    required_headers = ["课程代码", "课程名称"]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要列：{', '.join(missing)}")

    def _get_val(row, col_name):
        col = headers.get(col_name)
        if not col:
            return None
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        return str(val).strip()

    def _parse_int(val, default=0):
        if val is None or val == "":
            return default
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default

    def _parse_decimal(val, default=0):
        if val is None or val == "":
            return Decimal(str(default))
        try:
            return Decimal(str(val))
        except (ValueError, TypeError):
            return Decimal(str(default))

    def _parse_bool(val, default=True):
        if val is None or val == "":
            return default
        s = str(val).strip()
        if s in ("是", "Y", "y", "true", "True", "1"):
            return True
        if s in ("否", "N", "n", "false", "False", "0"):
            return False
        return default

    room_type_map = {
        "普通教室": "normal", "普通": "normal", "normal": "normal",
        "多媒体": "multimedia", "多媒体教室": "multimedia",
        "实验室": "lab", "lab": "lab",
        "计算机房": "computer", "机房": "computer",
        "艺术教室": "arts", "艺术": "arts",
    }

    success_count = 0
    fail_count = 0
    skip_count = 0
    failures = []

    for row in range(header_row + 1, ws.max_row + 1):
        course_code = _get_val(row, "课程代码")
        name = _get_val(row, "课程名称")

        if not course_code and not name:
            continue

        row_errors = []
        if not course_code:
            row_errors.append("课程代码不能为空")
        if not name:
            row_errors.append("课程名称不能为空")

        if course_code and db.query(Course).filter(Course.course_code == course_code).first():
            skip_count += 1
            continue

        # 院系匹配
        dept_name = _get_val(row, "开课院系")
        department_id = None
        if dept_name:
            dept = dept_by_name.get(dept_name)
            if dept:
                department_id = dept.id
            else:
                row_errors.append(f"院系「{dept_name}」不存在")

        if row_errors:
            fail_count += 1
            failures.append({"row": row, "name": name or course_code, "errors": row_errors})
            continue

        room_type_raw = _get_val(row, "所需教室类型") or "普通教室"
        required_room_type = room_type_map.get(room_type_raw, "normal")
        is_consecutive = _parse_bool(_get_val(row, "是否连堂"), True)

        course = Course(
            course_code=course_code,
            name=name,
            credit=_parse_decimal(_get_val(row, "学分"), 2),
            total_hours=_parse_int(_get_val(row, "总学时"), 32),
            hours_per_week=_parse_int(_get_val(row, "周学时"), 2),
            course_type=_get_val(row, "课程类型") or "必修",
            subject_type=_get_val(row, "课程性质") or "主课",
            required_room_type=required_room_type,
            is_consecutive=is_consecutive,
            consecutive_sections=_parse_int(_get_val(row, "连堂节数"), 2) if is_consecutive else 2,
            department_id=department_id,
            description=_get_val(row, "简介") or None,
        )
        db.add(course)
        success_count += 1

    if success_count > 0:
        db.commit()

    return {
        "message": f"导入完成：成功 {success_count} 条，跳过 {skip_count} 条，失败 {fail_count} 条",
        "success_count": success_count,
        "skip_count": skip_count,
        "fail_count": fail_count,
        "failures": failures[:50],
    }


@router.post("/import/teaching-tasks")
async def import_teaching_tasks(
    semester_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """批量导入教学任务（Excel）
    - 课程匹配：优先按课程编码，其次按课程名称
    - 教师匹配：按姓名，多个教师用逗号/分号分隔（第一位为主讲）
    - 班级匹配：按班级名称，多个班级用逗号/分号分隔
    """
    from openpyxl import load_workbook
    from app.models.teaching_task import TeachingTask, TaskTeacher, TaskClass
    from app.models.course import Course
    from app.models.teacher import Teacher
    from app.models.class_group import ClassGroup
    from app.models.semester import Semester

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件")

    # 检查学期
    semester = db.query(Semester).get(semester_id)
    if not semester:
        raise HTTPException(status_code=404, detail="学期不存在")

    # 加载已有数据用于匹配（缓存，减少查询）
    courses = db.query(Course).all()
    course_by_code = {c.course_code: c for c in courses if c.course_code}
    course_by_name = {c.name: c for c in courses}

    teachers = db.query(Teacher).filter(Teacher.status == 1).all()
    teacher_by_name = {}
    for t in teachers:
        teacher_by_name[t.name] = t

    classes = db.query(ClassGroup).filter(ClassGroup.status == 1).all()
    class_by_name = {c.name: c for c in classes}

    # 读取Excel
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    # 解析表头（跳过第一行标题）
    headers = {}
    header_row = 1
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col).value or "").strip()
        if val:
            headers[val] = col

    required_headers = ["课程名称"]
    missing = [h for h in required_headers if h not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要列：{', '.join(missing)}")

    # 辅助函数
    def _get_val(row, col_name):
        col = headers.get(col_name)
        if not col:
            return None
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        return str(val).strip()

    def _split_names(text):
        """按逗号/分号/顿号分隔姓名列表"""
        if not text:
            return []
        import re
        parts = re.split(r'[,，;；、\s]+', text.strip())
        return [p.strip() for p in parts if p.strip()]

    success_count = 0
    fail_count = 0
    failures = []
    created_tasks = []

    for row in range(header_row + 1, ws.max_row + 1):
        course_name = _get_val(row, "课程名称")
        course_code = _get_val(row, "课程编码")
        teacher_names_str = _get_val(row, "授课教师")
        class_names_str = _get_val(row, "授课班级")
        hours_str = _get_val(row, "周学时")
        priority_str = _get_val(row, "优先级")
        student_count_str = _get_val(row, "学生人数")
        task_code = _get_val(row, "任务编码")
        notes = _get_val(row, "备注")

        # 跳过空行
        if not course_name and not course_code and not teacher_names_str:
            continue

        row_errors = []

        # 匹配课程
        course = None
        if course_code and course_code in course_by_code:
            course = course_by_code[course_code]
        elif course_name and course_name in course_by_name:
            course = course_by_name[course_name]
        if not course:
            match_key = course_code or course_name
            row_errors.append(f"课程「{match_key}」不存在")

        # 匹配教师
        teacher_names = _split_names(teacher_names_str)
        matched_teachers = []
        for tname in teacher_names:
            t = teacher_by_name.get(tname)
            if t:
                matched_teachers.append(t)
            else:
                row_errors.append(f"教师「{tname}」不存在")

        # 匹配班级
        class_names = _split_names(class_names_str)
        matched_classes = []
        for cname in class_names:
            c = class_by_name.get(cname)
            if c:
                matched_classes.append(c)
            else:
                row_errors.append(f"班级「{cname}」不存在")

        if row_errors:
            fail_count += 1
            failures.append({
                "row": row,
                "course": course_name or course_code or "",
                "errors": row_errors,
            })
            continue

        # 解析数值字段
        hours_per_week = None
        if hours_str:
            try:
                hours_per_week = int(float(hours_str))
            except (ValueError, TypeError):
                hours_per_week = course.hours_per_week if course else None
        if not hours_per_week and course:
            hours_per_week = course.hours_per_week
        if not hours_per_week:
            hours_per_week = 2

        priority = 5
        if priority_str:
            try:
                priority = int(float(priority_str))
                priority = max(1, min(10, priority))
            except (ValueError, TypeError):
                priority = 5

        student_count = None
        if student_count_str:
            try:
                student_count = int(float(student_count_str))
            except (ValueError, TypeError):
                student_count = None

        # 创建教学任务
        task = TeachingTask(
            task_code=task_code or None,
            course_id=course.id,
            semester_id=semester_id,
            student_count=student_count,
            hours_per_week=hours_per_week,
            weeks="all",
            priority=priority,
            notes=notes or "",
            status=1,
        )
        if not task.task_code:
            task.task_code = f"TASK{semester_id:04d}{course.id:04d}{success_count + 1:03d}"

        db.add(task)
        db.flush()

        # 教师关联
        for idx, teacher in enumerate(matched_teachers):
            tt = TaskTeacher(task_id=task.id, teacher_id=teacher.id, is_main=(idx == 0))
            db.add(tt)

        # 班级关联
        for cls in matched_classes:
            tc = TaskClass(task_id=task.id, class_group_id=cls.id)
            db.add(tc)

        success_count += 1
        created_tasks.append(task)

    # 全部校验通过后统一提交
    if success_count > 0:
        db.commit()
        for task in created_tasks:
            db.refresh(task)

    return {
        "message": f"导入完成：成功 {success_count} 条，失败 {fail_count} 条",
        "success_count": success_count,
        "fail_count": fail_count,
        "failures": failures[:50],  # 最多返回50条失败详情
    }


# ---------- 导出 ----------
@router.get("/export/schedule")
async def export_schedule(
    semester_id: int,
    view_type: str = "class",
    db: Session = Depends(get_db),
):
    """导出课表Excel
    view_type: class（按班级）/ teacher（按教师）/ classroom（按教室）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from app.models.schedule import ScheduleEntry
    from app.models.class_group import ClassGroup
    from app.models.teacher import Teacher
    from app.models.classroom import Classroom
    from app.models.timeslot import TimeSlot
    from app.models.semester import Semester

    semester = db.query(Semester).get(semester_id)
    if not semester:
        raise HTTPException(status_code=404, detail="学期不存在")

    time_slots = db.query(TimeSlot).order_by(TimeSlot.section).all()
    sections = len(time_slots) or 10
    week_days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    entries = (
        db.query(ScheduleEntry)
        .filter(ScheduleEntry.semester_id == semester_id)
        .all()
    )

    wb = Workbook()
    wb.remove(wb.active)

    # 样式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    course_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def _fill_sheet(ws, item_entries_dict, days_count=5):
        """填充课表sheet"""
        # 标题
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_count + 1)
        ws.cell(row=1, column=1, value=f"{semester.name} 课表").font = title_font
        ws.cell(row=1, column=1).alignment = center_align

        # 表头
        ws.cell(row=3, column=1, value="节次\\星期").font = header_font
        ws.cell(row=3, column=1).fill = header_fill
        ws.cell(row=3, column=1).alignment = center_align
        ws.cell(row=3, column=1).border = thin_border

        for d in range(days_count):
            col = d + 2
            c = ws.cell(row=3, column=col, value=week_days[d])
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        # 节次行
        for sec in range(1, sections + 1):
            row = sec + 3
            time_label = time_slots[sec-1].start_time + "-" + time_slots[sec-1].end_time if sec <= len(time_slots) else ""
            cell = ws.cell(row=row, column=1, value=f"第{sec}节\n{time_label}")
            cell.alignment = center_align
            cell.border = thin_border
            for d in range(days_count):
                col = d + 2
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = thin_border

        # 填充课程
        for (day, sec_start), (course, teacher, room, span) in item_entries_dict.items():
            if day > days_count:
                continue
            row = sec_start + 3
            col = day + 1
            content_lines = [course]
            if teacher:
                content_lines.append(teacher)
            if room:
                content_lines.append(room)
            content = "\n".join(content_lines)
            cell = ws.cell(row=row, column=col, value=content)
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = course_fill
            if span > 1 and sec_start + span - 1 <= sections:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=row + span - 1, end_column=col,
                )

        # 列宽行高
        ws.column_dimensions['A'].width = 14
        for d in range(days_count):
            ws.column_dimensions[get_column_letter(d + 2)].width = 18
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[3].height = 22
        for sec in range(1, sections + 1):
            ws.row_dimensions[sec + 3].height = 40

    if view_type == "class":
        classes = db.query(ClassGroup).filter(ClassGroup.status == 1).order_by(ClassGroup.id).all()
        for cls in classes:
            class_entries = [e for e in entries if e.task and any(tc.class_group_id == cls.id for tc in e.task.task_classes)]
            cell_map = {}
            for e in class_entries:
                key = (e.day_of_week, e.section_start)
                course_name = e.task.course.name if e.task and e.task.course else ""
                teacher_names = ", ".join([tt.teacher.name for tt in e.task.task_teachers]) if e.task and e.task.task_teachers else ""
                room_name = f"{e.classroom.building}{e.classroom.room_number}" if e.classroom else ""
                span = e.section_end - e.section_start + 1
                cell_map[key] = (course_name, teacher_names, room_name, span)
            ws = wb.create_sheet(title=cls.name[:20])
            _fill_sheet(ws, cell_map)

    elif view_type == "teacher":
        teachers = db.query(Teacher).filter(Teacher.status == 1).order_by(Teacher.id).all()
        for teacher in teachers:
            t_entries = [e for e in entries if e.task and any(tt.teacher_id == teacher.id for tt in e.task.task_teachers)]
            cell_map = {}
            for e in t_entries:
                key = (e.day_of_week, e.section_start)
                course_name = e.task.course.name if e.task and e.task.course else ""
                class_names = ", ".join([tc.class_group.name for tc in e.task.task_classes]) if e.task and e.task.task_classes else ""
                room_name = f"{e.classroom.building}{e.classroom.room_number}" if e.classroom else ""
                span = e.section_end - e.section_start + 1
                cell_map[key] = (course_name, class_names, room_name, span)
            ws = wb.create_sheet(title=teacher.name[:20])
            _fill_sheet(ws, cell_map)

    elif view_type == "classroom":
        classrooms = db.query(Classroom).filter(Classroom.status == 1).order_by(Classroom.id).all()
        for room in classrooms:
            r_entries = [e for e in entries if e.classroom_id == room.id]
            cell_map = {}
            for e in r_entries:
                key = (e.day_of_week, e.section_start)
                course_name = e.task.course.name if e.task and e.task.course else ""
                teacher_names = ", ".join([tt.teacher.name for tt in e.task.task_teachers]) if e.task and e.task.task_teachers else ""
                class_names = ", ".join([tc.class_group.name for tc in e.task.task_classes]) if e.task and e.task.task_classes else ""
                span = e.section_end - e.section_start + 1
                cell_map[key] = (course_name, teacher_names, class_names, span)
            sheet_name = f"{room.building}{room.room_number}"[:20]
            ws = wb.create_sheet(title=sheet_name)
            _fill_sheet(ws, cell_map)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    view_label = {"class": "按班级", "teacher": "按教师", "classroom": "按教室"}.get(view_type, "")
    filename = f"{semester.name}_课表_{view_label}.xlsx"
    import urllib.parse
    encoded = urllib.parse.quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


def _generate_pdf_schedule(semester, time_slots, entries, view_type, db):
    """生成PDF课表，返回BytesIO"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    # 注册中文字体 - 尝试系统常见中文字体
    font_name = "Helvetica"
    font_bold = "Helvetica-Bold"
    font_paths = [
        "C:/Windows/Fonts/msyh.ttc",      # 微软雅黑
        "C:/Windows/Fonts/simhei.ttf",    # 黑体
        "C:/Windows/Fonts/simsun.ttc",    # 宋体
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/System/Library/Fonts/PingFang.ttc",
    ]
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ChineseFont", fp))
                font_name = "ChineseFont"
                font_bold = "ChineseFont"
                break
            except Exception:
                continue

    sections = len(time_slots) or 10
    week_days = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    days_count = 5

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleCN", parent=styles["Title"],
        fontName=font_bold, fontSize=18, leading=22, alignment=1,
    )
    header_style = ParagraphStyle(
        "HeaderCN", parent=styles["Normal"],
        fontName=font_bold, fontSize=10, leading=12, alignment=1, textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        "CellCN", parent=styles["Normal"],
        fontName=font_name, fontSize=8, leading=10, alignment=1,
    )
    section_style = ParagraphStyle(
        "SectionCN", parent=styles["Normal"],
        fontName=font_name, fontSize=8, leading=10, alignment=1,
    )

    story = []

    def build_cell_map(item_entries, get_course, get_meta1, get_meta2):
        """构造 {(day, sec_start): (course, meta1, meta2, span)}"""
        cell_map = {}
        for e in item_entries:
            key = (e.day_of_week, e.section_start)
            span = e.section_end - e.section_start + 1
            cell_map[key] = (get_course(e), get_meta1(e), get_meta2(e), span)
        return cell_map

    def add_sheet_page(title_text, cell_map):
        """添加一页课表"""
        story.append(Paragraph(title_text, title_style))
        story.append(Spacer(1, 6 * mm))

        # 构造表格数据
        data = [[""] + week_days[:days_count]]  # 表头
        for sec in range(1, sections + 1):
            time_label = ""
            if sec <= len(time_slots):
                time_label = f"<br/>{time_slots[sec-1].start_time}"
            row_text = Paragraph(f"第{sec}节{time_label}", section_style)
            row = [row_text]
            for d in range(1, days_count + 1):
                cell_text = ""
                if (d, sec) in cell_map:
                    course, m1, m2, span = cell_map[(d, sec)]
                    lines = [course]
                    if m1:
                        lines.append(m1)
                    if m2:
                        lines.append(m2)
                    cell_text = "<br/>".join(lines)
                row.append(Paragraph(cell_text, cell_style))
            data.append(row)

        # 计算列宽
        page_width = landscape(A4)[0] - 30 * mm
        col_widths = [18 * mm] + [(page_width - 18 * mm) / days_count] * days_count

        table = Table(data, colWidths=col_widths, repeatRows=1)

        # 合并单元格
        for (day, sec_start), (course, m1, m2, span) in cell_map.items():
            if span > 1 and day <= days_count and sec_start + span - 1 <= sections:
                table.setStyle(TableStyle([
                    ("SPAN", (day, sec_start), (day, sec_start + span - 1)),
                ]))

        # 表格样式
        table_style = TableStyle([
            # 表头
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            # 第一列（节次）
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#D9E2F3")),
            # 边框
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B4C7E7")),
            # 对齐
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            # 课程单元格背景
            ("BACKGROUND", (1, 1), (-1, -1), colors.HexColor("#FFF2CC")),
            # 行高
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
        table.setStyle(table_style)

        # 单独给有课程的单元格重新设置背景（覆盖默认）
        # 由于reportlab表格样式不好按cell条件设置，这里用空白单元格覆盖
        # 简化：整表用浅黄色，节次列用淡蓝色

        story.append(table)

    # 根据视图类型生成内容
    if view_type == "class":
        classes = db.query(ClassGroup).filter(ClassGroup.status == 1).order_by(ClassGroup.id).all()
        for i, cls in enumerate(classes):
            if i > 0:
                story.append(PageBreak())
            class_entries = [e for e in entries if e.task and any(tc.class_group_id == cls.id for tc in e.task.task_classes)]
            cell_map = build_cell_map(
                class_entries,
                lambda e: e.task.course.name if e.task and e.task.course else "",
                lambda e: ", ".join([tt.teacher.name for tt in e.task.task_teachers]) if e.task and e.task.task_teachers else "",
                lambda e: f"{e.classroom.building}{e.classroom.room_number}" if e.classroom else "",
            )
            add_sheet_page(f"{semester.name}  {cls.name} 课程表", cell_map)

    elif view_type == "teacher":
        teachers = db.query(Teacher).filter(Teacher.status == 1).order_by(Teacher.id).all()
        for i, teacher in enumerate(teachers):
            if i > 0:
                story.append(PageBreak())
            t_entries = [e for e in entries if e.task and any(tt.teacher_id == teacher.id for tt in e.task.task_teachers)]
            cell_map = build_cell_map(
                t_entries,
                lambda e: e.task.course.name if e.task and e.task.course else "",
                lambda e: ", ".join([tc.class_group.name for tc in e.task.task_classes]) if e.task and e.task.task_classes else "",
                lambda e: f"{e.classroom.building}{e.classroom.room_number}" if e.classroom else "",
            )
            add_sheet_page(f"{semester.name}  {teacher.name} 教师课表", cell_map)

    elif view_type == "classroom":
        classrooms = db.query(Classroom).filter(Classroom.status == 1).order_by(Classroom.id).all()
        for i, room in enumerate(classrooms):
            if i > 0:
                story.append(PageBreak())
            r_entries = [e for e in entries if e.classroom_id == room.id]
            cell_map = build_cell_map(
                r_entries,
                lambda e: e.task.course.name if e.task and e.task.course else "",
                lambda e: ", ".join([tt.teacher.name for tt in e.task.task_teachers]) if e.task and e.task.task_teachers else "",
                lambda e: ", ".join([tc.class_group.name for tc in e.task.task_classes]) if e.task and e.task.task_classes else "",
            )
            sheet_name = f"{room.building}{room.room_number}"
            add_sheet_page(f"{semester.name}  {sheet_name} 教室课表", cell_map)

    doc.build(story)
    output.seek(0)
    return output


@router.get("/export/schedule/pdf")
async def export_schedule_pdf(
    semester_id: int,
    view_type: str = "class",
    db: Session = Depends(get_db),
):
    """导出课表PDF
    view_type: class（按班级）/ teacher（按教师）/ classroom（按教室）
    """
    from app.models.schedule import ScheduleEntry
    from app.models.class_group import ClassGroup
    from app.models.teacher import Teacher
    from app.models.classroom import Classroom
    from app.models.timeslot import TimeSlot
    from app.models.semester import Semester
    from fastapi.responses import StreamingResponse
    import urllib.parse

    semester = db.query(Semester).get(semester_id)
    if not semester:
        raise HTTPException(status_code=404, detail="学期不存在")

    time_slots = db.query(TimeSlot).order_by(TimeSlot.section).all()
    entries = (
        db.query(ScheduleEntry)
        .filter(ScheduleEntry.semester_id == semester_id)
        .all()
    )

    output = _generate_pdf_schedule(semester, time_slots, entries, view_type, db)

    view_label = {"class": "按班级", "teacher": "按教师", "classroom": "按教室"}.get(view_type, "")
    filename = f"{semester.name}_课表_{view_label}.pdf"
    encoded = urllib.parse.quote(filename)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.get("/export/schedule/teacher/{teacher_id}")
async def export_teacher_schedule(teacher_id: int, semester_id: int, db: Session = Depends(get_db)):
    """导出单个教师课表"""
    # 复用上面的逻辑，传入指定教师
    from app.models.teacher import Teacher
    teacher = db.query(Teacher).get(teacher_id)
    if not teacher:
        raise HTTPException(status_code=404, detail="教师不存在")
    # 重定向到统一导出接口（此处简化，直接调用上面的逻辑）
    return await export_schedule(semester_id=semester_id, view_type="teacher", db=db)


@router.get("/export/schedule/class/{class_id}")
async def export_class_schedule(class_id: int, semester_id: int, db: Session = Depends(get_db)):
    """导出单个班级课表"""
    return await export_schedule(semester_id=semester_id, view_type="class", db=db)


@router.get("/import/template/{type}")
async def download_template(type: str):
    """下载导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from fastapi.responses import StreamingResponse
    import urllib.parse

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    def _add_headers(ws, headers, example=None):
        for col, (name, width, required) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=name + ("*" if required else ""))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        if example:
            for col, val in enumerate(example, 1):
                ws.cell(row=2, column=col, value=val)

    def _add_notes_sheet(wb, notes):
        ws2 = wb.create_sheet("填写说明")
        for i, note in enumerate(notes, 1):
            ws2.cell(row=i, column=1, value=note)
        ws2.column_dimensions['A'].width = 60

    if type == "teachers":
        ws = wb.active
        ws.title = "教师信息"
        headers = [
            ("工号", 14, True),
            ("姓名", 12, True),
            ("性别", 8, False),
            ("职称", 10, False),
            ("院系", 18, False),
            ("日最大课时", 12, False),
            ("周最大课时", 12, False),
            ("最大连堂", 10, False),
            ("午休需求", 10, False),
            ("电话", 14, False),
            ("邮箱", 22, False),
        ]
        example = ["T001", "张三", "男", "教授", "计算机学院", "6", "20", "4", "是", "13800138000", "zhangsan@school.edu.cn"]
        _add_headers(ws, headers, example)
        _add_notes_sheet(wb, [
            "教师导入模板填写说明",
            "",
            "1. 带 * 号的列为必填项",
            "2. 工号必须唯一，重复工号会被自动跳过",
            "3. 院系：需为系统中已存在的院系名称",
            "4. 职称：如教授、副教授、讲师、助教等",
            "5. 午休需求：填「是」或「否」，默认「是」",
            "6. 数据从第2行开始（第1行为表头，第2行为示例，导入前请删除示例行）",
        ])

    elif type == "classrooms":
        ws = wb.active
        ws.title = "教室信息"
        headers = [
            ("教室编号", 14, True),
            ("教学楼", 12, True),
            ("房间号", 10, True),
            ("容量", 10, False),
            ("类型", 14, False),
            ("校区", 10, False),
            ("设备", 30, False),
        ]
        example = ["A101", "教学楼A", "101", "60", "多媒体教室", "东校区", "投影仪、音响、白板"]
        _add_headers(ws, headers, example)
        _add_notes_sheet(wb, [
            "教室导入模板填写说明",
            "",
            "1. 带 * 号的列为必填项",
            "2. 教室编号必须唯一，重复编号会被自动跳过",
            "3. 教室类型可选值：普通教室 / 多媒体教室 / 实验室 / 计算机房 / 艺术教室",
            "4. 容量：不填默认为 50",
            "5. 设备：用顿号或逗号分隔多个设备",
            "6. 数据从第2行开始（第1行为表头，第2行为示例，导入前请删除示例行）",
        ])

    elif type == "courses":
        ws = wb.active
        ws.title = "课程信息"
        headers = [
            ("课程代码", 14, True),
            ("课程名称", 20, True),
            ("学分", 8, False),
            ("总学时", 10, False),
            ("周学时", 10, False),
            ("课程类型", 10, False),
            ("课程性质", 10, False),
            ("所需教室类型", 14, False),
            ("是否连堂", 10, False),
            ("连堂节数", 10, False),
            ("开课院系", 18, False),
            ("简介", 30, False),
        ]
        example = ["MATH101", "高等数学", "4", "64", "4", "必修", "主课", "普通教室", "是", "2", "数学学院", "公共基础课"]
        _add_headers(ws, headers, example)
        _add_notes_sheet(wb, [
            "课程导入模板填写说明",
            "",
            "1. 带 * 号的列为必填项",
            "2. 课程代码必须唯一，重复代码会被自动跳过",
            "3. 课程类型：必修 / 选修 / 公选",
            "4. 课程性质：主课 / 副课 / 实验 / 实践",
            "5. 所需教室类型：普通教室 / 多媒体教室 / 实验室 / 计算机房 / 艺术教室",
            "6. 是否连堂：填「是」或「否」，默认「是」",
            "7. 开课院系：需为系统中已存在的院系名称",
            "8. 数据从第2行开始（第1行为表头，第2行为示例，导入前请删除示例行）",
        ])

    elif type == "teaching-tasks":
        ws = wb.active

        # 示例行
        example = ["高等数学", "MATH101", "张三,李四", "计算机1班,计算机2班", "4", "8", "80", "", "公共基础课"]
        for col, val in enumerate(example, 1):
            ws.cell(row=2, column=col, value=val)

        _add_notes_sheet(wb, [
            "教学任务导入模板填写说明",
            "",
            "1. 带 * 号的列为必填项",
            "2. 课程匹配：优先按「课程编码」匹配，没有编码则按「课程名称」匹配",
            "3. 授课教师：多个教师用逗号分隔，第一位为主讲教师",
            "4. 授课班级：多个班级用逗号分隔（合班课）",
            "5. 周学时：不填则使用课程默认周学时",
            "6. 优先级：1-10，数字越大优先级越高，默认5",
            "7. 学生人数：不填则按班级人数自动计算",
            "8. 任务编码：不填则自动生成",
            "9. 数据从第2行开始（第1行为表头，第2行为示例，导入前请删除示例行）",
        ])

    else:
        ws = wb.active
        ws.cell(row=1, column=1, value=f"{type} 模板")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    name_map = {
        "teachers": "教师导入模板",
        "classrooms": "教室导入模板",
        "courses": "课程导入模板",
        "teaching-tasks": "教学任务导入模板",
    }
    filename = f"{name_map.get(type, type)}.xlsx"
    encoded = urllib.parse.quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
